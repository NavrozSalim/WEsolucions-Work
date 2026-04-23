"""
Catalog services: upload parsing, validation, and value normalization.
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from typing import Any, Generator, Iterable

from django.conf import settings
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction

from .models import CatalogUpload, CatalogUploadRow
from .marketplace_templates import INTERNAL_FIELDS, build_field_indices, validate_marketplace_headers
from .reverb_catalog import store_is_reverb, vendor_is_ebay
from stores.models import Store
from vendor.models import Vendor

logger = logging.getLogger(__name__)

# Backward-compatible alias (canonical list lives in marketplace_templates)
EXPECTED_COLUMNS = INTERNAL_FIELDS


CANONICAL_VENDOR_NAMES = (
    'AmazonUS', 'AmazonAU', 'EbayUS', 'EbayAU',
    'VevorAU', 'CostcoAU', 'HebUS',
)

_VENDOR_ALIAS_TO_CODE: dict[str, str] = {
    'amazonus': 'amazonus', 'amazonusa': 'amazonus', 'amazon us': 'amazonus',
    'amazon-us': 'amazonus', 'amazon_us': 'amazonus', 'amazon': 'amazonus',
    'amazonau': 'amazonau', 'amazon au': 'amazonau',
    'amazon-au': 'amazonau', 'amazon_au': 'amazonau',
    'ebayus': 'ebayus', 'ebay us': 'ebayus', 'ebay-us': 'ebayus',
    'ebay_us': 'ebayus', 'ebay': 'ebayus',
    'ebayau': 'ebayau', 'ebay au': 'ebayau', 'ebay-au': 'ebayau', 'ebay_au': 'ebayau',
    'vevorau': 'vevorau', 'vevor au': 'vevorau', 'vevor-au': 'vevorau',
    'vevor_au': 'vevorau', 'vevor': 'vevorau',
    'costcoau': 'costcoau', 'costco au': 'costcoau', 'costco-au': 'costcoau',
    'costco_au': 'costcoau', 'costco': 'costcoau',
    'hebus': 'hebus', 'heb us': 'hebus', 'heb-us': 'hebus',
    'heb_us': 'hebus', 'heb': 'hebus',
}


def resolve_canonical_vendor_code(raw: str) -> str | None:
    if not raw:
        return None
    key = str(raw).strip().lower()
    return _VENDOR_ALIAS_TO_CODE.get(key)


def _parse_xlsx_to_rows(file_obj) -> list[list[Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _parse_csv_bytes_to_rows(content: bytes | str) -> list[list[Any]]:
    if isinstance(content, bytes):
        content = content.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(content))
    return list(reader)


def parse_upload_file(file_obj, filename: str) -> list[list[Any]]:
    """Parse XLSX or CSV; return list of rows (first row = header). In-memory; prefer streaming ingest for large files."""
    name = (filename or '').lower()
    if name.endswith('.csv'):
        if hasattr(file_obj, 'read'):
            content = file_obj.read()
        else:
            with open(file_obj, 'rb') as f:
                content = f.read()
        return _parse_csv_bytes_to_rows(content)
    if name.endswith('.xlsx') or name.endswith('.xls'):
        return _parse_xlsx_to_rows(file_obj)
    raise ValueError("File must be CSV or XLSX")


def _store_raw(val: Any) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return s


def _normalize(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.upper() == 'N/A':
        return None
    return s


def get_catalog_upload_chunk_size() -> int:
    try:
        return max(200, int(getattr(settings, 'CATALOG_UPLOAD_CHUNK_SIZE', 1000)))
    except (TypeError, ValueError):
        return 1000


@dataclass
class RowIngestContext:
    store: Store
    indices: dict
    store_name_lower: str
    vendors_by_name: dict
    vendors_by_code: dict
    is_reverb: bool
    requires_fixed_inputs: bool


def _build_vendor_maps() -> tuple[dict, dict]:
    vendors_by_name = {v.name.lower(): v for v in Vendor.objects.all()}
    vendors_by_name.update({v.code.lower(): v for v in Vendor.objects.all()})
    vendors_by_code = {v.code.lower(): v for v in Vendor.objects.all()}
    return vendors_by_name, vendors_by_code


def _make_row_ingest_context(store: Store, header: list) -> RowIngestContext:
    from stores.models import StorePriceRangeMargin

    store = Store.objects.select_related('marketplace').get(pk=store.pk)
    indices = build_field_indices(header, store)
    header_err = validate_marketplace_headers(indices, store)
    if header_err:
        raise ValueError(header_err)
    requires_fixed = StorePriceRangeMargin.objects.filter(
        price_settings__store=store,
        margin_type='fixed',
    ).exists()
    vendors_by_name, vendors_by_code = _build_vendor_maps()
    return RowIngestContext(
        store=store,
        indices=indices,
        store_name_lower=store.name.lower(),
        vendors_by_name=vendors_by_name,
        vendors_by_code=vendors_by_code,
        is_reverb=store_is_reverb(store),
        requires_fixed_inputs=requires_fixed,
        is_ebay=None,
    )


def build_catalog_row_instance(
    upload: CatalogUpload,
    row_num: int,
    row: list,
    ctx: RowIngestContext,
) -> tuple[CatalogUploadRow | None, str | None]:
    """
    Build a single CatalogUploadRow or return an error string for invalid rows
    (same rules as the legacy per-row create).
    """
    def _val(row_in: list, col: str, raw: bool = True) -> str:
        i = ctx.indices.get(col)
        if i is None or i >= len(row_in):
            return ''
        v = row_in[i]
        return _store_raw(v) if raw else (_normalize(v) or '')

    store = ctx.store
    vendor_name_raw = _val(row, 'vendor name')
    vendor_id_raw = _val(row, 'vendor id')
    is_variation_raw = _val(row, 'is variation')
    variation_id_raw = _val(row, 'variation id')
    marketplace_name_raw = _val(row, 'marketplace name')
    store_name_raw = _val(row, 'store name')
    marketplace_parent_sku_raw = _val(row, 'marketplace parent sku')
    marketplace_child_sku_raw = _val(row, 'marketplace child sku')
    marketplace_id_raw = _val(row, 'marketplace id')
    vendor_sku_raw = _val(row, 'vendor sku')
    vendor_url_raw = _val(row, 'vendor url')
    action_raw = (_val(row, 'action') or 'Add').strip()
    pack_qty_raw = _val(row, 'pack qty')
    prep_fees_raw = _val(row, 'prep fees')
    shipping_fees_raw = _val(row, 'shipping fees')

    action_norm = action_raw.lower() if action_raw else 'add'
    if action_norm not in ('add', 'update', 'delete'):
        action_norm = 'add'

    if _normalize(vendor_name_raw) is None:
        return None, f"Row {row_num}: Vendor Name required (cannot be N/A or empty)"
    if _normalize(store_name_raw) is None:
        return None, f"Row {row_num}: Store Name required (cannot be N/A or empty)"

    if store_name_raw and store_name_raw.lower() != ctx.store_name_lower:
        # warned but row still created in legacy; keep same behavior
        pass

    vn = _normalize(vendor_name_raw)
    vendor = ctx.vendors_by_name.get(vn.lower()) if vn else None
    if not vendor and vn:
        canon = resolve_canonical_vendor_code(vn)
        if canon:
            vendor = ctx.vendors_by_code.get(canon)
    if not vendor and vn:
        return None, (
            f"Row {row_num}: Unknown vendor '{vendor_name_raw}'. "
            f"Use one of: {', '.join(CANONICAL_VENDOR_NAMES)}."
        )
    is_ebay_v = vendor_is_ebay(vendor, vendor_name_raw)

    if action_norm == 'add':
        if is_ebay_v:
            if _normalize(marketplace_parent_sku_raw) is None:
                return None, (
                    f"Row {row_num}: eBay vendor rows require Marketplace Parent SKU for Add "
                    f"(for marketplace listing / push; Child SKU, Marketplace ID, Vendor SKU may be N/A)"
                )
            if _normalize(vendor_url_raw) is None and _normalize(vendor_id_raw) is None:
                return None, (
                    f"Row {row_num}: eBay vendor rows require Vendor URL or Vendor ID "
                    f"(eBay item id for https://www.ebay.com/itm/...) for Add"
                )
        elif ctx.is_reverb:
            if _normalize(marketplace_parent_sku_raw) is None:
                return None, (
                    f"Row {row_num}: Reverb stores require SKU (or Marketplace Parent SKU) for Add "
                    f"(Reverb listing SKU; other marketplace columns may be N/A)"
                )
        else:
            sku_val = (
                _normalize(vendor_sku_raw)
                or _normalize(marketplace_child_sku_raw)
                or _normalize(vendor_id_raw)
                or _normalize(marketplace_parent_sku_raw)
            )
            if not sku_val:
                return None, (
                    f"Row {row_num}: Vendor SKU, Marketplace Child SKU, Vendor ID, or Marketplace Parent SKU required for Add"
                )
    if ctx.requires_fixed_inputs and action_norm in ('add', 'update'):
        if _normalize(pack_qty_raw) is None:
            return None, (
                f"Row {row_num}: Pack QTY required for {action_norm.title()} "
                f"(store uses a fixed pricing tier)"
            )
        if _normalize(prep_fees_raw) is None:
            return None, (
                f"Row {row_num}: Prep Fees required for {action_norm.title()} "
                f"(store uses a fixed pricing tier)"
            )
        if _normalize(shipping_fees_raw) is None:
            return None, (
                f"Row {row_num}: Shipping Fees required for {action_norm.title()} "
                f"(store uses a fixed pricing tier)"
            )
    elif action_norm == 'delete':
        id_val = (
            _normalize(marketplace_id_raw)
            or _normalize(vendor_sku_raw)
            or _normalize(vendor_id_raw)
            or _normalize(marketplace_child_sku_raw)
            or _normalize(marketplace_parent_sku_raw)
        )
        if not id_val:
            return None, (
                f"Row {row_num}: Delete requires Marketplace ID, Vendor SKU, Vendor ID, "
                f"Marketplace Child SKU, or Marketplace Parent SKU to find the product"
            )

    inst = CatalogUploadRow(
        catalog_upload=upload,
        row_number=row_num,
        vendor_name_raw=vendor_name_raw,
        vendor_id_raw=vendor_id_raw,
        is_variation_raw=is_variation_raw,
        variation_id_raw=variation_id_raw,
        marketplace_name_raw=marketplace_name_raw,
        store_name_raw=store_name_raw,
        marketplace_parent_sku_raw=marketplace_parent_sku_raw,
        marketplace_child_sku_raw=marketplace_child_sku_raw,
        marketplace_id_raw=marketplace_id_raw,
        vendor_sku_raw=vendor_sku_raw,
        vendor_url_raw=vendor_url_raw,
        action_raw=action_raw or 'Add',
        pack_qty_raw=pack_qty_raw,
        prep_fees_raw=prep_fees_raw,
        shipping_fees_raw=shipping_fees_raw,
        vendor=vendor,
        store=store,
    )
    return inst, None


def _error_summary_for_display(errors: list[str]) -> str | None:
    if not errors:
        return None
    tail = f" (+{len(errors) - 5} more)" if len(errors) > 5 else ""
    return "; ".join(errors[:5]) + tail


def _bulk_append_rows(
    upload: CatalogUpload,
    iter_rows: Generator[tuple[int, list], None, None] | Iterable[tuple[int, list]],
    ctx: RowIngestContext,
) -> tuple[int, list[str]]:
    """Insert rows in chunks. Returns (valid_row_count, error_messages)."""
    chunk_size = get_catalog_upload_chunk_size()
    errors: list[str] = []
    max_error_lines = 2000
    total_valid = 0
    buffer: list[CatalogUploadRow] = []

    for row_num, row in iter_rows:
        if len(errors) >= max_error_lines and not buffer:
            errors.append("Too many row errors; stopping error collection (ingest may continue for valid rows).")
        inst, err = build_catalog_row_instance(upload, row_num, row, ctx)
        if err:
            if len(errors) < max_error_lines:
                errors.append(err)
            continue
        if inst:
            sn = inst.store_name_raw
            if sn and sn.lower() != ctx.store_name_lower:
                if len(errors) < max_error_lines:
                    errors.append(
                        f"Row {row_num}: Store Name '{sn}' does not match upload store "
                        f"'{ctx.store.name}' (row still created)"
                    )
            buffer.append(inst)
            if len(buffer) >= chunk_size:
                with transaction.atomic():
                    CatalogUploadRow.objects.bulk_create(buffer, batch_size=500)
                total_valid += len(buffer)
                buffer = []
    if buffer:
        with transaction.atomic():
            CatalogUploadRow.objects.bulk_create(buffer, batch_size=500)
        total_valid += len(buffer)
    return total_valid, errors


def ingest_stored_catalog_file(upload_id) -> dict:
    """
    Celery: parse source_file, chunked bulk_create. Idempotent: clears existing rows for upload on start.
    """
    upload = (
        CatalogUpload.objects.select_related('store', 'store__marketplace')
        .get(id=upload_id)
    )
    if not upload.source_file:
        return {'error': 'no_source_file', 'upload_id': str(upload_id)}
    store = upload.store
    fn = (upload.original_filename or '').lower()

    # Ensure clean slate for safe retries
    with transaction.atomic():
        upload.rows.all().delete()
        upload.total_rows = 0
        upload.processed_rows = 0
        upload.save(update_fields=['total_rows', 'processed_rows'])

    try:
        if fn.endswith('.csv'):
            with upload.source_file.open('rb') as raw:
                text = io.TextIOWrapper(raw, encoding='utf-8-sig', newline='')
                reader = csv.reader(text)
                try:
                    header = next(reader)
                except StopIteration:
                    raise ValueError("File is empty") from None
                try:
                    ctx = _make_row_ingest_context(store, header)
                except ValueError as e:
                    upload.status = CatalogUpload.Status.FAILED
                    upload.error_summary = str(e)[:2000]
                    upload.save(update_fields=['status', 'error_summary'])
                    return {'error': str(e), 'upload_id': str(upload_id)}

                def row_data() -> Generator[tuple[int, list], None, None]:
                    for row_num, row in enumerate(reader, start=2):
                        yield row_num, row

                total_valid, row_errors = _bulk_append_rows(upload, row_data(), ctx)
        elif fn.endswith('.xlsx') or fn.endswith('.xls'):
            import openpyxl

            path = upload.source_file.path
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                wb.close()
                raise ValueError("File is empty")
            header = [str(c) if c is not None else '' for c in header_row]
            try:
                ctx = _make_row_ingest_context(store, header)
            except ValueError as e:
                wb.close()
                upload.status = CatalogUpload.Status.FAILED
                upload.error_summary = str(e)[:2000]
                upload.save(update_fields=['status', 'error_summary'])
                return {'error': str(e), 'upload_id': str(upload_id)}

            def row_data() -> Generator[tuple[int, list], None, None]:
                try:
                    row_num = 2
                    for row in it:
                        row_list = list(row) if row is not None else []
                        yield row_num, row_list
                        row_num += 1
                finally:
                    wb.close()

            total_valid, row_errors = _bulk_append_rows(upload, row_data(), ctx)
        else:
            raise ValueError("File must be CSV or XLSX")
    except Exception as e:
        upload.status = CatalogUpload.Status.FAILED
        upload.error_summary = str(e)[:2000]
        upload.save(update_fields=['status', 'error_summary'])
        logger.exception("Catalog ingest failed for upload_id=%s", upload_id)
        return {'error': str(e), 'upload_id': str(upload_id)}

    upload.total_rows = total_valid
    if row_errors:
        upload.error_summary = _error_summary_for_display(row_errors)
    else:
        upload.error_summary = None
    upload.status = CatalogUpload.Status.VALIDATED
    update_fields = ['total_rows', 'status', 'error_summary', 'processed_rows']
    if upload.source_file:
        upload.source_file.delete(save=False)
        upload.source_file = None
        update_fields.append('source_file')
    upload.save(update_fields=update_fields)
    return {
        'upload_id': str(upload_id),
        'status': upload.status,
        'total_rows': total_valid,
        'ingest_error_count': len(row_errors),
    }


def create_upload_file_and_queue(
    *,
    user,
    store: Store,
    file_obj: UploadedFile,
    filename: str,
) -> tuple[CatalogUpload | None, str | None]:
    """
    Save file, create INGESTING upload, caller enqueues catalog_ingest_upload_file_task.
    Returns (upload, None) or (None, error message).
    """
    safe_name = (filename or getattr(file_obj, 'name', 'upload.csv') or 'upload.csv')[:255]
    upload = CatalogUpload(
        user=user,
        store=store,
        original_filename=safe_name,
        total_rows=0,
        processed_rows=0,
        status=CatalogUpload.Status.INGESTING,
    )
    upload.save()
    upload.source_file.save(safe_name, file_obj, save=True)
    return upload, None


def validate_and_create_upload(
    *,
    user,
    store: Store,
    file_obj,
    filename: str,
) -> tuple[CatalogUpload, list[str]]:
    """
    Legacy synchronous path: parse full file in memory, chunked bulk insert.
    Prefer create_upload_file_and_queue + Celery for large files.
    """
    errors: list[str] = []
    try:
        rows = parse_upload_file(file_obj, filename)
    except Exception as e:
        return None, [str(e)]
    if not rows:
        return None, ["File is empty"]
    header = rows[0]
    store = Store.objects.select_related('marketplace').get(pk=store.pk)
    try:
        ctx = _make_row_ingest_context(store, header)
    except ValueError as e:
        return None, [str(e)]

    with transaction.atomic():
        upload = CatalogUpload.objects.create(
            user=user,
            store=store,
            original_filename=filename,
            total_rows=0,
            processed_rows=0,
            status=CatalogUpload.Status.INGESTING,
        )

    data_iter = ((row_num, r) for row_num, r in enumerate(rows[1:], start=2))
    total_valid, row_errors = _bulk_append_rows(upload, data_iter, ctx)
    errors.extend(row_errors)
    upload.total_rows = total_valid
    upload.error_summary = _error_summary_for_display(errors) if errors else None
    upload.status = CatalogUpload.Status.VALIDATED
    upload.save(update_fields=['total_rows', 'processed_rows', 'status', 'error_summary'])
    return upload, errors
