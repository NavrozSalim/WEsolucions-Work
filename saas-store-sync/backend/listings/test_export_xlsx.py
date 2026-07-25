"""Tests for managed-store Excel exports."""
from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
import io

from marketplace.models import Marketplace
from stores.models import Store

from listings.export_xlsx import build_orders_xlsx
from listings.models import MarketplaceOrder, OrderShipment, OrderStatus


class OrdersExportXlsxTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(username="ex", email="ex@example.com", password="pw")
        self.mp, _ = Marketplace.objects.get_or_create(
            code="reverb",
            defaults={"name": "Reverb"},
        )
        self.store = Store.objects.create(
            user=self.user,
            name="IQ",
            region="USA",
            marketplace=self.mp,
            management_mode="full_store",
        )

    def test_orders_export_includes_customer_address_and_tracking(self):
        order = MarketplaceOrder.objects.create(
            user=self.user,
            store=self.store,
            external_order_key="26084333",
            invoice_number="26084333",
            status=OrderStatus.SHIPPING_COMPLETE,
            shipping_status="submitted",
            total_amount_cents=1747,
            environment="production",
            customer_info_json={
                "firstName": "Christopher",
                "lastName": "Emerson",
                "email": "chris@example.com",
                "phone": "555-0100",
            },
            line_items_json=[
                {
                    "title": "Cable Adapter",
                    "quantity": 1,
                    "priceCents": 1499,
                    "sku": "CAB-1",
                }
            ],
            raw_response_json={
                "status": "shipped",
                "shipping": {
                    "status": "shipped",
                    "carrier": "USPS",
                    "trackingNumber": "9400TRACK",
                    "trackingUrl": "https://track.example/9400TRACK",
                    "method": "Priority",
                    "address": {
                        "line1": "12 Main St",
                        "line2": "Apt 4",
                        "city": "Austin",
                        "state": "TX",
                        "postcode": "78701",
                        "country": "US",
                    },
                },
                "totals": {
                    "subtotalCents": 1499,
                    "shippingCents": 248,
                    "taxCents": 0,
                    "totalCents": 1747,
                    "currency": "USD",
                },
            },
        )
        OrderShipment.objects.create(
            order=order,
            carrier="USPS",
            tracking_number="APP-TRACK-1",
            tracking_url="https://app.example/track/1",
            status="submitted",
        )

        content = build_orders_xlsx([order], self.store)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertIn("Email", headers)
        self.assertIn("Ship address 1", headers)
        self.assertIn("Tracking (marketplace)", headers)
        self.assertIn("Tracking (submitted)", headers)
        self.assertIn("Vendor URL", headers)

        row = [c.value for c in ws[2]]
        data = dict(zip(headers, row))
        self.assertEqual(data["Order ID"], "26084333")
        self.assertEqual(data["Buyer Name"], "Christopher Emerson")
        self.assertEqual(data["Email"], "chris@example.com")
        self.assertEqual(data["Phone"], "555-0100")
        self.assertEqual(data["Ship address 1"], "12 Main St")
        self.assertEqual(data["Ship city"], "Austin")
        self.assertEqual(data["Ship postcode"], "78701")
        self.assertEqual(data["Title"], "Cable Adapter")
        self.assertEqual(data["SKU"], "CAB-1")
        self.assertEqual(data["Quantity"], 1)
        self.assertEqual(data["Carrier (marketplace)"], "USPS")
        self.assertEqual(data["Tracking (marketplace)"], "9400TRACK")
        self.assertEqual(data["Tracking (submitted)"], "APP-TRACK-1")
        self.assertEqual(data["Order total"], 17.47)
