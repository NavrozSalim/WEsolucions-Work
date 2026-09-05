"""Parse uploaded CSV / Excel listing templates into normalized row dicts.

Uses stdlib csv + openpyxl (no pandas dependency in this project).
Supports Lasoo, Reverb, MyDeal, Etsy, and Bunnings templates (by store marketplace).

Optional columns are labeled with `` (Optional)`` in the template header.
Import still accepts headers with or without that suffix.
"""
import csv
import io
import json
import re

from stores.credentials import marketplace_kind

# Maps human template headers -> internal field names (Lasoo + Reverb + MyDeal + Bunnings).
COLUMN_MAP = {
    "action": "action",
    "product key": "product_key",
    "parent sku": "product_key",
    "parent_sku": "product_key",
    "variant key": "variant_key",
    "title": "title",
    "description": "description",
    "brand": "brand",
    "make": "make",
    "model": "model",
    "finish": "finish",
    "year": "year",
    "category": "category",
    "category uuid": "category_uuid",
    "category id": "category",
    "taxonomy id": "taxonomy_id",
    "taxonomy_id": "taxonomy_id",
    "who made": "who_made",
    "who_made": "who_made",
    "when made": "when_made",
    "when_made": "when_made",
    "shipping profile id": "shipping_profile_id",
    "shipping_profile_id": "shipping_profile_id",
    "readiness state id": "readiness_state_id",
    "readiness_state_id": "readiness_state_id",
    "mydeal category id": "category",
    "condition": "condition",
    "condition uuid": "condition_uuid",
    "sku": "sku",
    "barcode": "barcode",
    "upc": "barcode",
    "gtin": "gtin",
    "mpn": "mpn",
    "upc does not apply": "upc_does_not_apply",
    "options": "options",
    "option": "options",
    "variant options": "options",
    "option 1 name": "option_1_name",
    "option1 name": "option_1_name",
    "option 1 value": "option_1_value",
    "option1 value": "option_1_value",
    "option 2 name": "option_2_name",
    "option2 name": "option_2_name",
    "option 2 value": "option_2_value",
    "option2 value": "option_2_value",
    "option 3 name": "option_3_name",
    "option3 name": "option_3_name",
    "option 3 value": "option_3_value",
    "option3 value": "option_3_value",
    "option 4 name": "option_4_name",
    "option4 name": "option_4_name",
    "option 4 value": "option_4_value",
    "option4 value": "option_4_value",
    "variation img url": "variation_image_url",
    "variation image url": "variation_image_url",
    "variant image url": "variation_image_url",
    "variant img url": "variation_image_url",
    "currency": "currency",
    "price": "sale_price",
    "vendor url": "vendor_url",
    "vendor_url": "vendor_url",
    "source url": "vendor_url",
    "source link": "vendor_url",
    "vendor link": "vendor_url",
    "vendor id": "vendor_id",
    "vendor_id": "vendor_id",
    "vendor name": "vendor_name",
    "vendor_name": "vendor_name",
    "source vendor": "vendor_name",
    "marketplace name": "marketplace_name",
    "marketplace_name": "marketplace_name",
    "marketplace": "marketplace_name",
    "store name": "store_name",
    "store_name": "store_name",
    "store": "store_name",
    "image urls": "image_urls",
    "image url": "image_urls",
    "photos": "image_urls",
    "photo urls": "image_urls",
    "inventory": "inventory",
    "infinite quantity": "infinite_quantity",
    "product unlimited": "infinite_quantity",
    "original price": "original_price",
    "rrp": "original_price",
    "sale price": "sale_price",
    "status": "publish_status",
    "publish status": "publish_status",
    "listing status": "publish_status",
    "free_shipping": "free_shipping",
    "free shipping": "free_shipping",
    # MyDeal / WMP ProductGroup extras
    "tags": "tags",
    "specifications": "specifications",
    "weight": "weight",
    "weight unit": "weight_unit",
    "length": "length",
    "height": "height",
    "width": "width",
    "dimension unit": "dimension_unit",
    "shipping cost category": "shipping_cost_category",
    "shipping cost standard": "shipping_cost_standard",
    "custom freight scheme id": "custom_freight_scheme_id",
    "is direct import": "is_direct_import",
    "max days for delivery": "max_days_for_delivery",
    "delivery time": "delivery_time",
    "has 48 hours dispatch": "has_48_hours_dispatch",
    "logistic class": "logistic_class",
    "logistic-class": "logistic_class",
    "leadtime to ship": "leadtime_to_ship",
    "lead time to ship": "leadtime_to_ship",
    "leadtime-to-ship": "leadtime_to_ship",
    "category code": "category",
    "variant-group-code": "product_key",
    "variant group code": "product_key",
    "variant_group_code": "product_key",
    "category attributes json": "attributes",
    "category attributes": "attributes",
}


def _parse_attributes_blob(value) -> dict:
    if isinstance(value, dict):
        return {
            str(k).strip(): str(v).strip()
            for k, v in value.items()
            if str(k or "").strip() and v not in (None, "")
        }
    text = str(value or "").strip()
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {
        str(k).strip(): str(v).strip()
        for k, v in parsed.items()
        if str(k or "").strip() and v not in (None, "")
    }


def _is_attribute_header(header: str) -> bool:
    text = str(header or "").strip().lower()
    return text.startswith("attribute_") or text.startswith("pdb_") or "attribute_pdb" in text

_OPTIONAL_SUFFIX_RE = re.compile(r"\s*\(optional\)\s*$", re.I)


def _canonical_header(header: str) -> str:
    """Lowercase header and strip a trailing ``(Optional)`` marker."""
    text = str(header or "").strip().lower()
    return _OPTIONAL_SUFFIX_RE.sub("", text).strip()


LASOO_TEMPLATE_HEADERS = [
    "Vendor Name (Optional)",
    "Vendor URL (Optional)",
    "Vendor ID (Optional)",
    "Marketplace Name (Optional)",
    "Store Name (Optional)",
    "Action",
    "Parent SKU",
    "SKU",
    "Option 1 Name (Optional)",
    "Option 1 Value (Optional)",
    "Option 2 Name (Optional)",
    "Option 2 Value (Optional)",
    "Option 3 Name (Optional)",
    "Option 3 Value (Optional)",
    "Option 4 Name (Optional)",
    "Option 4 Value (Optional)",
    "Variation Img URL (Optional)",
    "Title",
    "Description",
    "Brand",
    "Category (Optional)",
    "Barcode (Optional)",
    "Image URLs",
    "Inventory",
    "Infinite Quantity (Optional)",
    "Original Price (Optional)",
    "Sale Price",
]

LASOO_EXPORT_FIELDS = [
    ("vendor_name", "Vendor Name (Optional)"),
    ("vendor_url", "Vendor URL (Optional)"),
    ("vendor_id", "Vendor ID (Optional)"),
    ("marketplace_name", "Marketplace Name (Optional)"),
    ("store_name", "Store Name (Optional)"),
    ("action", "Action"),
    ("product_key", "Parent SKU"),
    ("sku", "SKU"),
    ("option_1_name", "Option 1 Name (Optional)"),
    ("option_1_value", "Option 1 Value (Optional)"),
    ("option_2_name", "Option 2 Name (Optional)"),
    ("option_2_value", "Option 2 Value (Optional)"),
    ("option_3_name", "Option 3 Name (Optional)"),
    ("option_3_value", "Option 3 Value (Optional)"),
    ("option_4_name", "Option 4 Name (Optional)"),
    ("option_4_value", "Option 4 Value (Optional)"),
    ("variation_image_url", "Variation Img URL (Optional)"),
    ("title", "Title"),
    ("description", "Description"),
    ("brand", "Brand"),
    ("category", "Category (Optional)"),
    ("barcode", "Barcode (Optional)"),
    ("image_urls", "Image URLs"),
    ("inventory", "Inventory"),
    ("infinite_quantity", "Infinite Quantity (Optional)"),
    ("original_price", "Original Price (Optional)"),
    ("sale_price", "Sale Price"),
]

REVERB_EXPORT_FIELDS = [
    ("vendor_name", "Vendor Name (Optional)"),
    ("vendor_url", "Vendor URL (Optional)"),
    ("marketplace_name", "Marketplace Name (Optional)"),
    ("store_name", "Store Name (Optional)"),
    ("action", "Action"),
    ("sku", "SKU"),
    ("title", "Title"),
    ("make", "Make"),
    ("model", "Model"),
    ("description", "Description"),
    ("finish", "Finish (Optional)"),
    ("year", "Year (Optional)"),
    ("condition", "Condition"),
    ("category", "Category"),
    ("sale_price", "Price"),
    ("currency", "Currency (Optional)"),
    ("inventory", "Inventory"),
    ("barcode", "UPC (Optional)"),
    ("upc_does_not_apply", "UPC Does Not Apply (Optional)"),
    ("image_urls", "Photo URLs"),
    ("publish_status", "status (Optional)"),
    ("free_shipping", "free_shipping (Optional)"),
]

REVERB_TEMPLATE_HEADERS = [
    "Vendor Name (Optional)",
    "Vendor URL (Optional)",
    "Marketplace Name (Optional)",
    "Store Name (Optional)",
    "Action",
    "SKU",
    "Title",
    "Make",
    "Model",
    "Description",
    "Finish (Optional)",
    "Year (Optional)",
    "Condition",
    "Category",
    "Price",
    "Currency (Optional)",
    "Inventory",
    "UPC (Optional)",
    "UPC Does Not Apply (Optional)",
    "Photo URLs",
    "status (Optional)",
    "free_shipping (Optional)",
]

ETSY_EXPORT_FIELDS = [
    ("vendor_name", "Vendor Name (Optional)"),
    ("vendor_url", "Vendor URL (Optional)"),
    ("marketplace_name", "Marketplace Name (Optional)"),
    ("store_name", "Store Name (Optional)"),
    ("action", "Action"),
    ("sku", "SKU"),
    ("title", "Title"),
    ("description", "Description"),
    ("taxonomy_id", "Taxonomy ID"),
    ("who_made", "Who Made"),
    ("when_made", "When Made"),
    ("sale_price", "Price"),
    ("inventory", "Inventory"),
    ("image_urls", "Photo URLs"),
    ("shipping_profile_id", "Shipping Profile ID (Optional)"),
    ("readiness_state_id", "Readiness State ID (Optional)"),
    ("publish_status", "status (Optional)"),
]

ETSY_TEMPLATE_HEADERS = [
    "Vendor Name (Optional)",
    "Vendor URL (Optional)",
    "Marketplace Name (Optional)",
    "Store Name (Optional)",
    "Action",
    "SKU",
    "Title",
    "Description",
    "Taxonomy ID",
    "Who Made",
    "When Made",
    "Price",
    "Inventory",
    "Photo URLs",
    "Shipping Profile ID (Optional)",
    "Readiness State ID (Optional)",
    "status (Optional)",
]

MYDEAL_TEMPLATE_HEADERS = [
    "Vendor Name (Optional)",
    "Vendor URL (Optional)",
    "Vendor ID (Optional)",
    "Marketplace Name (Optional)",
    "Store Name (Optional)",
    "Action",
    "Parent SKU",
    "SKU",
    "Title",
    "Description",
    "Brand (Optional)",
    "Tags (Optional)",
    "Specifications (Optional)",
    "Condition (Optional)",
    "Category",
    "GTIN (Optional)",
    "MPN (Optional)",
    "Image URLs",
    "Inventory",
    "Product Unlimited (Optional)",
    "Price",
    "RRP (Optional)",
    "Weight (Optional)",
    "Weight Unit (Optional)",
    "Length (Optional)",
    "Height (Optional)",
    "Width (Optional)",
    "Dimension Unit (Optional)",
    "Shipping Cost Category",
    "Shipping Cost Standard (Optional)",
    "Custom Freight Scheme ID (Optional)",
    "Is Direct Import",
    "Max Days For Delivery",
    "Delivery Time",
    "Has 48 Hours Dispatch (Optional)",
    "Option 1 Name (Optional)",
    "Option 1 Value (Optional)",
    "Option 2 Name (Optional)",
    "Option 2 Value (Optional)",
    "Option 3 Name (Optional)",
    "Option 3 Value (Optional)",
]

MYDEAL_EXPORT_FIELDS = [
    ("vendor_name", "Vendor Name (Optional)"),
    ("vendor_url", "Vendor URL (Optional)"),
    ("vendor_id", "Vendor ID (Optional)"),
    ("marketplace_name", "Marketplace Name (Optional)"),
    ("store_name", "Store Name (Optional)"),
    ("action", "Action"),
    ("product_key", "Parent SKU"),
    ("sku", "SKU"),
    ("title", "Title"),
    ("description", "Description"),
    ("brand", "Brand (Optional)"),
    ("tags", "Tags (Optional)"),
    ("specifications", "Specifications (Optional)"),
    ("condition", "Condition (Optional)"),
    ("category", "Category"),
    ("gtin", "GTIN (Optional)"),
    ("mpn", "MPN (Optional)"),
    ("image_urls", "Image URLs"),
    ("inventory", "Inventory"),
    ("infinite_quantity", "Product Unlimited (Optional)"),
    ("sale_price", "Price"),
    ("original_price", "RRP (Optional)"),
    ("weight", "Weight (Optional)"),
    ("weight_unit", "Weight Unit (Optional)"),
    ("length", "Length (Optional)"),
    ("height", "Height (Optional)"),
    ("width", "Width (Optional)"),
    ("dimension_unit", "Dimension Unit (Optional)"),
    ("shipping_cost_category", "Shipping Cost Category"),
    ("shipping_cost_standard", "Shipping Cost Standard (Optional)"),
    ("custom_freight_scheme_id", "Custom Freight Scheme ID (Optional)"),
    ("is_direct_import", "Is Direct Import"),
    ("max_days_for_delivery", "Max Days For Delivery"),
    ("delivery_time", "Delivery Time"),
    ("has_48_hours_dispatch", "Has 48 Hours Dispatch (Optional)"),
    ("option_1_name", "Option 1 Name (Optional)"),
    ("option_1_value", "Option 1 Value (Optional)"),
    ("option_2_name", "Option 2 Name (Optional)"),
    ("option_2_value", "Option 2 Value (Optional)"),
    ("option_3_name", "Option 3 Name (Optional)"),
    ("option_3_value", "Option 3 Value (Optional)"),
]

BUNNINGS_TEMPLATE_HEADERS = [
    "Vendor Name (Optional)",
    "Vendor URL (Optional)",
    "Vendor ID (Optional)",
    "Marketplace Name (Optional)",
    "Store Name (Optional)",
    "Action",
    "Parent SKU",
    "SKU",
    "Option 1 Name (Optional)",
    "Option 1 Value (Optional)",
    "Option 2 Name (Optional)",
    "Option 2 Value (Optional)",
    "Option 3 Name (Optional)",
    "Option 3 Value (Optional)",
    "Option 4 Name (Optional)",
    "Option 4 Value (Optional)",
    "Variation Img URL (Optional)",
    "Title",
    "Description",
    "Brand",
    "Category",
    "GTIN (Optional)",
    "MPN (Optional)",
    "Image URLs",
    "Inventory",
    "Price",
    "RRP (Optional)",
    "Logistic Class",
    "Leadtime To Ship (Optional)",
    "Weight (Optional)",
    "Weight Unit (Optional)",
    "Length (Optional)",
    "Height (Optional)",
    "Width (Optional)",
    "Dimension Unit (Optional)",
    "Category Attributes JSON (Optional)",
]

BUNNINGS_EXPORT_FIELDS = [
    ("vendor_name", "Vendor Name (Optional)"),
    ("vendor_url", "Vendor URL (Optional)"),
    ("vendor_id", "Vendor ID (Optional)"),
    ("marketplace_name", "Marketplace Name (Optional)"),
    ("store_name", "Store Name (Optional)"),
    ("action", "Action"),
    ("product_key", "Parent SKU"),
    ("sku", "SKU"),
    ("option_1_name", "Option 1 Name (Optional)"),
    ("option_1_value", "Option 1 Value (Optional)"),
    ("option_2_name", "Option 2 Name (Optional)"),
    ("option_2_value", "Option 2 Value (Optional)"),
    ("option_3_name", "Option 3 Name (Optional)"),
    ("option_3_value", "Option 3 Value (Optional)"),
    ("option_4_name", "Option 4 Name (Optional)"),
    ("option_4_value", "Option 4 Value (Optional)"),
    ("variation_image_url", "Variation Img URL (Optional)"),
    ("title", "Title"),
    ("description", "Description"),
    ("brand", "Brand"),
    ("category", "Category"),
    ("gtin", "GTIN (Optional)"),
    ("mpn", "MPN (Optional)"),
    ("image_urls", "Image URLs"),
    ("inventory", "Inventory"),
    ("sale_price", "Price"),
    ("original_price", "RRP (Optional)"),
    ("logistic_class", "Logistic Class"),
    ("leadtime_to_ship", "Leadtime To Ship (Optional)"),
    ("weight", "Weight (Optional)"),
    ("weight_unit", "Weight Unit (Optional)"),
    ("length", "Length (Optional)"),
    ("height", "Height (Optional)"),
    ("width", "Width (Optional)"),
    ("dimension_unit", "Dimension Unit (Optional)"),
    ("attributes", "Category Attributes JSON (Optional)"),
]

TEMPLATE_HEADERS = LASOO_TEMPLATE_HEADERS
DELETE_TEMPLATE_HEADERS = ["Action", "SKU"]
VALID_ACTIONS = {"create", "mapped", "delete"}
_TRUE_VALUES = {"true", "1", "yes", "y", "t"}

_HEADER_MARKERS = {
    "action",
    "sku",
    "product key",
    "parent sku",
    "variant key",
    "vendor name",
    "marketplace name",
    "store name",
    "title",
    "sale price",
    "original price",
    "price",
    "image urls",
    "photo urls",
}


def _is_reverb_store(store) -> bool:
    return marketplace_kind(getattr(store, "marketplace", None)) == "reverb"


def _is_mydeal_store(store) -> bool:
    return marketplace_kind(getattr(store, "marketplace", None)) == "mydeal"


def _is_etsy_store(store) -> bool:
    return marketplace_kind(getattr(store, "marketplace", None)) == "etsy"


def _is_bunnings_store(store) -> bool:
    return marketplace_kind(getattr(store, "marketplace", None)) == "bunnings"


def _coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in _TRUE_VALUES


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_action(raw: str) -> str:
    """Accept plain Create/Mapped/Delete or instructional cells like 'Create - …'."""
    action = str(raw or "").strip().lower()
    if action in VALID_ACTIONS:
        return action
    for valid in ("create", "mapped", "delete"):
        if action.startswith(valid):
            return valid
    return ""


def _header_score(cells: list[str]) -> int:
    score = 0
    for cell in cells:
        key = _canonical_header(cell)
        if key in _HEADER_MARKERS or key in COLUMN_MAP:
            score += 1
    return score


def _pick_header_row(raw_rows: list[tuple]) -> tuple[list[str], list[tuple]]:
    """Return (headers, data_rows). Skips banner rows used in Excel templates."""
    if not raw_rows:
        return [], []
    best_idx, best_score = 0, -1
    scan_limit = min(5, len(raw_rows))
    for idx in range(scan_limit):
        cells = [_cell_text(c) for c in raw_rows[idx]]
        score = _header_score(cells)
        if score > best_score:
            best_idx, best_score = idx, score
    if best_score < 2:
        best_idx = 0
    headers = [_cell_text(c) for c in raw_rows[best_idx]]
    return headers, raw_rows[best_idx + 1 :]


def _records_from_header_and_rows(headers: list[str], data_rows: list[tuple]) -> list[dict]:
    """Build row dicts. Duplicate headers keep the last non-empty value."""
    records = []
    for row in data_rows:
        record: dict[str, str] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = _cell_text(row[i]) if i < len(row) else ""
            if header in record and not value:
                continue
            record[header] = value
        records.append(record)
    return records


def _read_xlsx_rows(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        raw_rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        headers, data_rows = _pick_header_row(raw_rows)
        return _records_from_header_and_rows(headers, data_rows)
    finally:
        wb.close()


def _read_csv_rows(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text.splitlines()[:5]
    if len(sample) >= 2:
        scored = []
        for i, line in enumerate(sample):
            try:
                cells = next(csv.reader([line]))
            except Exception:  # noqa: BLE001
                cells = [line]
            scored.append((_header_score([_cell_text(c) for c in cells]), i))
        scored.sort(reverse=True)
        best_score, best_idx = scored[0]
        if best_score >= 2 and best_idx > 0:
            lines = text.splitlines(keepends=True)
            text = "".join(lines[best_idx:])
    reader = csv.DictReader(io.StringIO(text))
    return [
        {(k or "").strip(): (v or "").strip() for k, v in raw.items()}
        for raw in reader
    ]


def parse_upload(filename: str, content: bytes) -> list[dict]:
    """Return a list of row dicts (1-based 'row_number' included; row 1 = header)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        records = _read_xlsx_rows(content)
    else:
        records = _read_csv_rows(content)

    rows = []
    for idx, raw in enumerate(records, start=2):
        normalized = {}
        extra_attrs = {}
        for header, value in raw.items():
            key = COLUMN_MAP.get(_canonical_header(header))
            if key:
                existing = normalized.get(key)
                if existing and not str(value).strip():
                    continue
                if existing and str(value).strip():
                    normalized[key] = str(value).strip()
                    continue
                normalized[key] = str(value).strip()
            elif _is_attribute_header(header) and str(value).strip():
                extra_attrs[str(header).strip()] = str(value).strip()
        if not any(str(v).strip() for v in normalized.values() if not isinstance(v, bool)):
            # Allow bool-only rows? skip empty
            if not any(normalized.values()) and not extra_attrs:
                continue
        normalized["infinite_quantity"] = _coerce_bool(normalized.get("infinite_quantity", ""))
        normalized["upc_does_not_apply"] = _coerce_bool(normalized.get("upc_does_not_apply", ""))
        normalized["is_direct_import"] = _coerce_bool(normalized.get("is_direct_import", ""))
        normalized["has_48_hours_dispatch"] = _coerce_bool(normalized.get("has_48_hours_dispatch", ""))
        fs_raw = normalized.get("free_shipping")
        if fs_raw is None or str(fs_raw).strip() == "":
            normalized["free_shipping"] = True
        else:
            normalized["free_shipping"] = _coerce_bool(fs_raw)
        ps = str(normalized.get("publish_status") or "").strip().lower()
        if ps in ("live", "published", "publish"):
            normalized["publish_status"] = "live"
        else:
            normalized["publish_status"] = "draft"
        if normalized.get("make") and not normalized.get("brand"):
            normalized["brand"] = normalized["make"]
        if normalized.get("condition") and not normalized.get("condition_uuid"):
            normalized["condition_uuid"] = normalized["condition"]
        if normalized.get("category_uuid") and not normalized.get("category"):
            normalized["category"] = normalized["category_uuid"]
        if normalized.get("category") and not normalized.get("category_uuid"):
            normalized["category_uuid"] = normalized["category"]
        if normalized.get("sale_price") and not normalized.get("original_price"):
            # Don't overwrite RRP for MyDeal when only Price is set — only fill if blank.
            if not str(normalized.get("original_price") or "").strip():
                normalized["original_price"] = normalized["sale_price"]
        sku = str(normalized.get("sku") or normalized.get("variant_key") or "").strip()
        parent = str(normalized.get("product_key") or "").strip()
        if not sku:
            sku = parent
        if sku:
            normalized["sku"] = sku
            if not str(normalized.get("variant_key") or "").strip():
                normalized["variant_key"] = sku
        # Default MyDeal required shipping/delivery fields when blank (template sample fills them).
        if not str(normalized.get("shipping_cost_category") or "").strip():
            # Leave blank for non-MyDeal rows; publish layer has its own defaults.
            pass
        normalized["action"] = _normalize_action(normalized.get("action", ""))
        merged_attrs = _parse_attributes_blob(normalized.get("attributes"))
        merged_attrs.update(extra_attrs)
        if merged_attrs:
            normalized["attributes"] = merged_attrs
        elif "attributes" in normalized:
            normalized.pop("attributes", None)
        normalized["row_number"] = idx
        rows.append(normalized)
    return rows


def _marketplace_label(store) -> str:
    mp = getattr(store, "marketplace", None)
    return (getattr(mp, "name", None) or getattr(mp, "code", None) or "").strip()


def build_template_csv(action: str = "create", store=None) -> str:
    """Template CSV for the given action. Headers mark optional columns."""
    action = (action or "create").strip().lower()
    if action == "delete":
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=DELETE_TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerow({"Action": "Delete", "SKU": "AMH-EXAMPLE-001"})
        return out.getvalue()

    store_name = (getattr(store, "name", None) or "").strip()
    marketplace_name = _marketplace_label(store) if store is not None else ""

    if _is_reverb_store(store):
        sample = {
            "Vendor Name (Optional)": "Amazon US",
            "Vendor URL (Optional)": "https://www.amazon.com/dp/EXAMPLE",
            "Marketplace Name (Optional)": marketplace_name or "Reverb",
            "Store Name (Optional)": store_name,
            "Action": "Mapped" if action == "mapped" else "Create",
            "SKU": "AMH-EXAMPLE-001",
            "Title": "Example Guitar Pedal",
            "Make": "Unbranded",
            "Model": "EXAMPLE-001",
            "Description": "Great pedal in excellent condition.",
            "Finish (Optional)": "",
            "Year (Optional)": "",
            "Condition": "Brand New",
            "Category": "Accessories / Cables",
            "Price": "49.99",
            "Currency (Optional)": "USD",
            "Inventory": "1",
            "UPC (Optional)": "",
            "UPC Does Not Apply (Optional)": "true",
            "Photo URLs": "https://example.com/photo1.jpg|https://example.com/photo2.jpg",
            "status (Optional)": "draft",
            "free_shipping (Optional)": "TRUE",
        }
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=REVERB_TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerow(sample)
        return out.getvalue()

    if _is_etsy_store(store):
        sample = {
            "Vendor Name (Optional)": "Amazon US",
            "Vendor URL (Optional)": "https://www.amazon.com/dp/EXAMPLE",
            "Marketplace Name (Optional)": marketplace_name or "Etsy",
            "Store Name (Optional)": store_name,
            "Action": "Mapped" if action == "mapped" else "Create",
            "SKU": "ETSY-EXAMPLE-001",
            "Title": "Handmade Example Pendant",
            "Description": "Example Etsy listing description.",
            "Taxonomy ID": "1",
            "Who Made": "someone_else",
            "When Made": "2020_2024",
            "Price": "24.99",
            "Inventory": "3",
            "Photo URLs": "https://example.com/photo1.jpg",
            "Shipping Profile ID (Optional)": "",
            "Readiness State ID (Optional)": "",
            "status (Optional)": "draft",
        }
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=ETSY_TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerow(sample)
        return out.getvalue()

    if _is_mydeal_store(store):
        def _row(*, sku, parent_sku, option_name, option_value, title_suffix=""):
            return {
                "Vendor Name (Optional)": "Amazon AU",
                "Vendor URL (Optional)": "https://www.amazon.com.au/dp/EXAMPLE",
                "Vendor ID (Optional)": "",
                "Marketplace Name (Optional)": marketplace_name or "MyDeal",
                "Store Name (Optional)": store_name,
                "Action": "Mapped" if action == "mapped" else "Create",
                "Parent SKU": parent_sku,
                "SKU": sku,
                "Title": f"Example Polo Shirt{title_suffix}",
                "Description": "Example product description for MyDeal / WMP. Same Parent SKU groups size variants.",
                "Brand (Optional)": "ExampleBrand",
                "Tags (Optional)": "apparel,polo",
                "Specifications (Optional)": "",
                "Condition (Optional)": "New",
                "Category": "3213",
                "GTIN (Optional)": "",
                "MPN (Optional)": "",
                "Image URLs": "https://example.com/photo1.jpg|https://example.com/photo2.jpg",
                "Inventory": "5",
                "Product Unlimited (Optional)": "false",
                "Price": "79.99",
                "RRP (Optional)": "99.99",
                "Weight (Optional)": "0.4",
                "Weight Unit (Optional)": "kg",
                "Length (Optional)": "30",
                "Height (Optional)": "2",
                "Width (Optional)": "20",
                "Dimension Unit (Optional)": "cm",
                "Shipping Cost Category": "Flat",
                "Shipping Cost Standard (Optional)": "0",
                "Custom Freight Scheme ID (Optional)": "",
                "Is Direct Import": "false",
                "Max Days For Delivery": "10",
                "Delivery Time": "5-10 business days",
                "Has 48 Hours Dispatch (Optional)": "false",
                "Option 1 Name (Optional)": option_name,
                "Option 1 Value (Optional)": option_value,
                "Option 2 Name (Optional)": "",
                "Option 2 Value (Optional)": "",
                "Option 3 Name (Optional)": "",
                "Option 3 Value (Optional)": "",
            }

        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=MYDEAL_TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerow(_row(
            sku="MD-EXAMPLE-001-S",
            parent_sku="MD-EXAMPLE-001",
            option_name="Size",
            option_value="Small",
        ))
        writer.writerow(_row(
            sku="MD-EXAMPLE-001-M",
            parent_sku="MD-EXAMPLE-001",
            option_name="Size",
            option_value="Medium",
        ))
        return out.getvalue()

    if _is_bunnings_store(store):
        sample = {
            "Vendor Name (Optional)": "Amazon AU",
            "Vendor URL (Optional)": "https://www.amazon.com.au/dp/EXAMPLE",
            "Vendor ID (Optional)": "",
            "Marketplace Name (Optional)": marketplace_name or "Bunnings",
            "Store Name (Optional)": store_name,
            "Action": "Mapped" if action == "mapped" else "Create",
            "Parent SKU": "BN-EXAMPLE-001",
            "SKU": "BN-EXAMPLE-001-M",
            "Option 1 Name (Optional)": "Size",
            "Option 1 Value (Optional)": "M",
            "Option 2 Name (Optional)": "Colour",
            "Option 2 Value (Optional)": "Red",
            "Option 3 Name (Optional)": "",
            "Option 3 Value (Optional)": "",
            "Option 4 Name (Optional)": "",
            "Option 4 Value (Optional)": "",
            "Variation Img URL (Optional)": "https://example.com/photo-red-m.jpg",
            "Title": "Example Power Drill",
            "Description": "Example product description for Bunnings Marketplace.",
            "Brand": "ExampleBrand",
            "Category": "HIERARCHY_CODE",
            "GTIN (Optional)": "9300000000001",
            "MPN (Optional)": "",
            "Image URLs": "https://example.com/photo1.jpg|https://example.com/photo2.jpg",
            "Inventory": "5",
            "Price": "79.99",
            "RRP (Optional)": "99.99",
            "Logistic Class": "SMALL",
            "Leadtime To Ship (Optional)": "2",
            "Weight (Optional)": "2.5",
            "Weight Unit (Optional)": "kg",
            "Length (Optional)": "30",
            "Height (Optional)": "20",
            "Width (Optional)": "15",
            "Dimension Unit (Optional)": "cm",
            "Category Attributes JSON (Optional)": "",
        }
        out = io.StringIO()
        writer = csv.DictWriter(out, fieldnames=BUNNINGS_TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerow(sample)
        return out.getvalue()

    sample_black = {
        "Vendor Name (Optional)": "Nora Inventory",
        "Vendor URL (Optional)": "https://www.example-vendor.com/product/jdxty-xl-b",
        "Vendor ID (Optional)": "8FNZ100-DL-G1",
        "Marketplace Name (Optional)": marketplace_name or "Lasoo",
        "Store Name (Optional)": store_name,
        "Action": "Mapped" if action == "mapped" else "Create",
        "Parent SKU": "JDXTY",
        "SKU": "JDXTY-XL-B",
        "Option 1 Name (Optional)": "Size",
        "Option 1 Value (Optional)": "XL",
        "Option 2 Name (Optional)": "Color",
        "Option 2 Value (Optional)": "Blue",
        "Option 3 Name (Optional)": "",
        "Option 3 Value (Optional)": "",
        "Option 4 Name (Optional)": "",
        "Option 4 Value (Optional)": "",
        "Variation Img URL (Optional)": "https://img.example.com/jdxty-xl-blue.jpg",
        "Title": "Example Product — XL Blue",
        "Description": "Same Parent SKU groups size/colour variants on Lasoo.",
        "Brand": "MyBrand",
        "Category (Optional)": "Apparel > T-Shirts",
        "Barcode (Optional)": "123456789012",
        "Image URLs": "https://img.example.com/a.jpg|https://img.example.com/b.jpg",
        "Inventory": "10",
        "Infinite Quantity (Optional)": "false",
        "Original Price (Optional)": "29.99",
        "Sale Price": "24.99",
    }
    sample_red = {
        **sample_black,
        "SKU": "JDXTY-S-R",
        "Option 1 Name (Optional)": "Size",
        "Option 1 Value (Optional)": "S",
        "Option 2 Name (Optional)": "Color",
        "Option 2 Value (Optional)": "Red",
        "Variation Img URL (Optional)": "https://img.example.com/jdxty-s-red.jpg",
        "Title": "Example Product — S Red",
        "Vendor URL (Optional)": "https://www.example-vendor.com/product/jdxty-s-r",
        "Vendor ID (Optional)": "8FNZ100-DL-G2",
        "Barcode (Optional)": "123456789013",
    }
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=LASOO_TEMPLATE_HEADERS, lineterminator="\n")
    writer.writeheader()
    writer.writerow(sample_black)
    writer.writerow(sample_red)
    return out.getvalue()


def snapshot_row_fields(row: dict) -> dict:
    """Copy template-relevant fields for later export (JSON-safe scalars)."""
    out = {}
    for key, value in (row or {}).items():
        if key in ("row_number", "errors", "valid", "imported", "fields"):
            continue
        if isinstance(value, bool):
            out[key] = "true" if value else "false"
        elif value is None:
            out[key] = ""
        elif isinstance(value, (dict, list)):
            out[key] = json.dumps(value)
        else:
            out[key] = str(value).strip() if not isinstance(value, (int, float)) else value
    action = str(out.get("action") or "").strip().lower()
    if action in VALID_ACTIONS:
        out["action"] = action.capitalize()
    return out


def export_field_specs(store=None) -> list[tuple[str, str]]:
    """Return (internal_key, header) pairs for the store's marketplace template."""
    if _is_reverb_store(store):
        return list(REVERB_EXPORT_FIELDS)
    if _is_etsy_store(store):
        return list(ETSY_EXPORT_FIELDS)
    if _is_mydeal_store(store):
        return list(MYDEAL_EXPORT_FIELDS)
    if _is_bunnings_store(store):
        return list(BUNNINGS_EXPORT_FIELDS)
    return list(LASOO_EXPORT_FIELDS)
