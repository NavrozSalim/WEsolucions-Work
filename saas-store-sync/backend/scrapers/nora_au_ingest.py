"""Nora AU inventory Excel parser.

Reads the ``Export inventory`` sheet and builds a Vendor ID → stock map using
the pivot / suffix-normalization rules agreed for Nora Inventory:

1. Sum ``Available inventory`` for exact duplicate ``BarCode`` values.
2. Strip trailing ``-G1/-G2/-G3/-V1/-V2/-V3`` suffixes.
3. If a suffix was stripped, append ``-G1`` (canonical Vendor ID).
   Barcodes with no suffix stay unchanged.
4. Sum again by the canonical Vendor ID.

Column layout (1-based): BarCode = 1 (index 0), Available inventory = 11 (index 10).
"""
from __future__ import annotations

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import BinaryIO

logger = logging.getLogger("scrapers.nora_au_ingest")

SHEET_NAME = "Export inventory"
BARCODE_COL = 0  # Excel column 1
AVAILABLE_INVENTORY_COL = 10  # Excel column 11

_SUFFIX_RE = re.compile(r"-(G[123]|V[123])$", re.IGNORECASE)
_CANONICAL_SUFFIX = "-G1"


def normalize_nora_vendor_id(barcode: str) -> str:
    """Apply Nora Vendor ID cleaning to a single barcode string."""
    raw = (barcode or "").strip()
    if not raw:
        return ""
    if _SUFFIX_RE.search(raw):
        base = _SUFFIX_RE.sub("", raw)
        return f"{base}{_CANONICAL_SUFFIX}" if base else raw
    return raw


def _to_int_stock(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def _sum_by_key(pairs: list[tuple[str, int]]) -> dict[str, int]:
    out: dict[str, int] = defaultdict(int)
    for key, stock in pairs:
        if not key:
            continue
        out[key] += stock
    return dict(out)


def build_nora_stock_map_from_rows(rows) -> dict[str, int]:
    """Build Vendor ID → stock from an iterable of row tuples/lists."""
    first_pass: list[tuple[str, int]] = []
    for row in rows:
        if not row:
            continue
        cell = row[BARCODE_COL] if len(row) > BARCODE_COL else ""
        barcode = str(cell or "").strip()
        if not barcode or barcode.lower() in ("barcode", "bar code"):
            continue
        stock = _to_int_stock(
            row[AVAILABLE_INVENTORY_COL] if len(row) > AVAILABLE_INVENTORY_COL else 0
        )
        first_pass.append((barcode, stock))

    # Pivot 1: exact barcode duplicates summed
    exact = _sum_by_key(first_pass)

    # Normalize suffixes → canonical Vendor ID, then pivot 2
    second_pass = [(normalize_nora_vendor_id(bc), stock) for bc, stock in exact.items()]
    return _sum_by_key(second_pass)


def load_nora_stock_map_from_path(path: str | Path) -> dict[str, int]:
    """Load and parse a Nora Excel file from disk."""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active
            logger.warning(
                "Nora file %s has no %r sheet; using active sheet %r",
                path.name,
                SHEET_NAME,
                ws.title,
            )
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return {}
    # Skip header row when it looks like the export header
    header = rows[0]
    start = 1
    if header and str(header[0] or "").strip().lower() in ("barcode", "bar code"):
        start = 1
    else:
        start = 0
    return build_nora_stock_map_from_rows(rows[start:])


def load_nora_stock_map_from_file(file_obj: BinaryIO) -> dict[str, int]:
    """Load and parse a Nora Excel file-like object (Django FieldFile / UploadedFile)."""
    import tempfile
    import os

    suffix = ".xlsx"
    name = getattr(file_obj, "name", "") or ""
    if str(name).lower().endswith(".xls"):
        suffix = ".xls"
    tmp = tempfile.NamedTemporaryFile(prefix="nora_au_", suffix=suffix, delete=False)
    try:
        if hasattr(file_obj, "open"):
            try:
                file_obj.open("rb")
            except Exception:
                pass
        if hasattr(file_obj, "seek"):
            try:
                file_obj.seek(0)
            except Exception:
                pass
        for chunk in file_obj.chunks() if hasattr(file_obj, "chunks") else [file_obj.read()]:
            if chunk:
                tmp.write(chunk)
        tmp.close()
        return load_nora_stock_map_from_path(tmp.name)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def is_nora_vendor_code(code: str | None) -> bool:
    c = (code or "").strip().lower()
    return c in ("noraau", "nora", "nora_au", "nora-au", "norainventory", "nora inventory")
