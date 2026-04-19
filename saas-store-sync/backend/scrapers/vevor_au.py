"""
VevorAU "scraper" — reads the public S3 XLSX feed and looks up price/stock per SKU.

Unlike Amazon/eBay, Vevor AU publishes a live catalog feed on S3. Instead of
scraping product pages (which gets rate-limited / Cloudflare-blocked), we:

  1. Download https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/563/vevor-563.xlsx
  2. Read columns A (SKU), G (Price), I (Inventory) — 0-based indices 0, 6, 8.
  3. Build a SKU -> {price, stock} lookup.
  4. Update ``VendorPrice`` rows for matching products.

The dispatcher (``scrapers.get_price_and_stock``) now treats vevor.com.au URLs
as ingest-only so the catalog task falls back to the latest VendorPrice just
like HEB and Costco. The actual refresh runs via the Celery task
``catalog.tasks.run_vevor_au_ingest`` (see catalog/tasks.py).
"""
from __future__ import annotations

import logging
import os
import re
import tempfile
from decimal import Decimal
from typing import Iterable

import requests

logger = logging.getLogger("scrapers.vevor_au")

VEVOR_AU_FEED_URL = os.getenv(
    "VEVOR_AU_FEED_URL",
    "https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/563/vevor-563.xlsx",
)


def _ingest_only_result() -> dict:
    """Sentinel for scrapers.get_price_and_stock — force VendorPrice fallback."""
    return {
        "price": None,
        "inventory": None,
        "title": None,
        "error_code": "vevor_ingest_only",
        "error_message": (
            "Vevor AU is fed from the public S3 catalog XLSX, not scraped per-URL. "
            "Run catalog.tasks.run_vevor_au_ingest to refresh VendorPrice."
        ),
    }


def clean_id(value) -> str:
    """Normalize a SKU cell: strip whitespace, zero-width chars, trailing .0 from Excel floats."""
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\u200b", "").replace("\xa0", " ").strip()
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


_COMPACT_RE = re.compile(r"[^A-Za-z0-9]+")


def compact_id(sku: str) -> str:
    """Return an alphanumeric-only lowercase key for fuzzy SKU matches."""
    if not sku:
        return ""
    return _COMPACT_RE.sub("", sku).lower()


def parse_price_value(value) -> float:
    """Accept 'USD 12.34', '12,34', '12.34', 12.34 — return float (0.0 on failure)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        try:
            return float(value)
        except Exception:
            return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("\xa0", " ").strip()
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except Exception:
        return 0.0


def parse_inventory_value(value) -> int:
    """Accept '10', '10 units', '', None — return int >= 0."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        try:
            n = int(value)
            return max(0, n)
        except Exception:
            return 0
    s = str(value).strip()
    if not s:
        return 0
    m = re.search(r"\d+", s.replace(",", ""))
    if not m:
        return 0
    try:
        return max(0, int(m.group(0)))
    except Exception:
        return 0


def round_precise(val: float, digits: int = 2) -> float:
    try:
        return float(Decimal(str(val)).quantize(Decimal("1." + ("0" * digits))))
    except Exception:
        return round(float(val), digits)


def load_veror_via_excel_positions(path: str) -> tuple[dict, dict, int]:
    """
    Positional read of the Vevor AU feed XLSX.

    Columns (0-based): A=0 SKU, G=6 Price, I=8 Inventory.

    Returns ``(lookup, lookup_compact, pos_rows_scanned)``:
    - ``lookup``: {sku: {'Posted Price': float, 'Posted Inventory': int}}
    - ``lookup_compact``: same by ``compact_id(sku)`` for fuzzy match fallback.
    - ``pos_rows_scanned``: number of data rows scanned.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        lookup: dict[str, dict] = {}
        lookup_compact: dict[str, dict] = {}
        pos_rows = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                continue
            if row is None or len(row) < 9:
                continue
            pos_rows += 1
            sku = clean_id(row[0])
            if not sku:
                continue
            price = round_precise(parse_price_value(row[6]), 2)
            stock = parse_inventory_value(row[8])
            entry = {"Posted Price": price, "Posted Inventory": int(stock)}
            lookup[sku] = entry
            ckey = compact_id(sku)
            if ckey:
                lookup_compact[ckey] = entry
        return lookup, lookup_compact, pos_rows
    finally:
        wb.close()


def fetch_vevor_feed(url: str = VEVOR_AU_FEED_URL, timeout: int = 60) -> str:
    """Download the Vevor AU XLSX to a temp file and return its path."""
    resp = requests.get(url, timeout=timeout, stream=True)
    resp.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(prefix="vevor_au_", suffix=".xlsx", delete=False)
    try:
        for chunk in resp.iter_content(chunk_size=65536):
            if chunk:
                tmp.write(chunk)
    finally:
        tmp.close()
    return tmp.name


def lookup_sku(lookup: dict, lookup_compact: dict, sku: str) -> dict | None:
    """Try exact SKU first, then alphanumeric-only fuzzy key."""
    if not sku:
        return None
    exact = lookup.get(sku)
    if exact:
        return exact
    ckey = compact_id(sku)
    if not ckey:
        return None
    return lookup_compact.get(ckey)


def iter_vevor_entries(lookup: dict) -> Iterable[tuple[str, dict]]:
    return lookup.items()


__all__ = [
    "VEVOR_AU_FEED_URL",
    "_ingest_only_result",
    "clean_id",
    "compact_id",
    "parse_price_value",
    "parse_inventory_value",
    "round_precise",
    "load_veror_via_excel_positions",
    "fetch_vevor_feed",
    "lookup_sku",
]
