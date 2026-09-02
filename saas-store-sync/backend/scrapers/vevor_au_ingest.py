"""
Vevor AU ingest — reads the public S3 XLSX feed and looks up price/stock per SKU.

Unlike Amazon/eBay, Vevor AU publishes a live catalog feed on S3. Instead of
scraping product pages (which gets rate-limited / Cloudflare-blocked), we:

  1. Download https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/563/vevor-563.xlsx
  2. Resolve columns from the header row: ``SKU``, ``after coupon price``
     (preferred over ``MAP``) and ``Inventory quantity``. If the header is
     unrecognizable we fall back to the legacy positional layout A=SKU,
     G=Price, I=Inventory (0-based 0, 6, 8) that the feed used before Vevor
     expanded it to a full catalog export.
  3. Build a SKU -> {price, stock} lookup.
  4. Update ``VendorPrice`` rows for matching products.

The dispatcher (``scrapers.get_price_and_stock``) now treats vevor.com.au URLs
as ingest-only so the catalog task falls back to the latest VendorPrice just
like HEB and Costco. The actual refresh runs via the Celery task
``catalog.tasks.run_vevor_au_ingest`` (see catalog/tasks.py).
"""
from __future__ import annotations

import logging
import os
import re
import tempfile
from decimal import Decimal
from typing import Iterable

import requests

logger = logging.getLogger("scrapers.vevor_au_ingest")

DEFAULT_VEVOR_AU_FEED_URL = (
    "https://ads-feed.s3.us-west-2.amazonaws.com/ads/business/563/vevor-563.xlsx"
)


def resolve_vevor_au_feed_url(raw: str | None = None) -> str:
    """Return feed URL; treat unset/blank env as the public default (Docker may pass '')."""
    if raw is None:
        raw = os.getenv("VEVOR_AU_FEED_URL")
    url = (raw or "").strip()
    return url or DEFAULT_VEVOR_AU_FEED_URL


VEVOR_AU_FEED_URL = resolve_vevor_au_feed_url()


def _ingest_only_result() -> dict:
    """Sentinel for scrapers.get_price_and_stock — force VendorPrice fallback."""
    return {
        "price": None,
        "inventory": None,
        "title": None,
        "error_code": "vevor_ingest_only",
        "error_message": (
            "Vevor AU is fed from the public S3 catalog XLSX, not scraped per-URL. "
            "Run catalog.tasks.run_vevor_au_ingest to refresh VendorPrice."
        ),
    }


def is_vevor_vendor_code(code: str | None) -> bool:
    """True for Vevor / Vevor AU vendor codes (vevor, vevorau, vevor_au, …)."""
    c = (code or "").strip().lower().replace("-", "").replace("_", "").replace(" ", "")
    return c == "vevor" or c.startswith("vevor")


def is_vevor_product_url(url: str | None) -> bool:
    u = (url or "").strip().lower()
    return "vevor.com.au" in u or "vevor.au" in u


def normalize_vevor_product_url(url: str | None) -> str:
    """Strip scheme, www, query, and trailing slash for Product-link matching."""
    s = (url or "").strip()
    if not s:
        return ""
    s = s.split("#", 1)[0].split("?", 1)[0].rstrip("/").lower()
    s = re.sub(r"^https?://(www\.)?", "", s)
    return s


_P_SUFFIX_RE = re.compile(r"(?:^|[-_])p[_-]?([A-Za-z0-9]{4,})$", re.I)


def vevor_identity_candidates(
    *,
    vendor_id: str = "",
    sku: str = "",
    variant_key: str = "",
    product_key: str = "",
    vendor_url: str = "",
) -> list[str]:
    """Ordered product-ID guesses to match against the Vevor feed SKU column."""
    from urllib.parse import parse_qs, unquote, urlparse

    keys: list[str] = []
    seen: set[str] = set()

    def add(val) -> None:
        s = clean_id(val)
        if not s:
            return
        marker = s.lower()
        if marker in seen:
            return
        seen.add(marker)
        keys.append(s)

    add(vendor_id)
    add(sku)
    add(variant_key)
    add(product_key)
    url = (vendor_url or "").strip()
    if url:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        for qk in ("sku", "SKU", "id", "product_id", "productId"):
            for val in qs.get(qk, []):
                add(unquote(val or ""))
        path = unquote(parsed.path or "").rstrip("/")
        last = path.split("/")[-1] if path else ""
        last = re.sub(r"\.(html?|php)$", "", last, flags=re.I)
        if last:
            m = _P_SUFFIX_RE.search(last)
            if m:
                add(m.group(1))
            add(last)
    return keys


def lookup_vevor_price_stock(
    lookup: dict,
    lookup_compact: dict,
    lookup_by_url: dict | None = None,
    *,
    vendor_id: str = "",
    sku: str = "",
    variant_key: str = "",
    product_key: str = "",
    vendor_url: str = "",
) -> dict | None:
    """Find a feed row by Product link, then by Vendor ID / SKU / URL tokens."""
    url = (vendor_url or "").strip()
    if url and lookup_by_url:
        hit = lookup_by_url.get(normalize_vevor_product_url(url))
        if hit:
            return hit
    for key in vevor_identity_candidates(
        vendor_id=vendor_id,
        sku=sku,
        variant_key=variant_key,
        product_key=product_key,
        vendor_url=vendor_url,
    ):
        hit = lookup_sku(lookup, lookup_compact, key)
        if hit:
            return hit
    return None


def _product_link_column_index(header_row) -> int | None:
    for idx, cell in enumerate(header_row or ()):
        name = _normalize_header(cell)
        if name in ("product link", "product url"):
            return idx
    return None


def load_vevor_feed_lookups() -> dict:
    """Download and parse the live Vevor AU XLSX once per scrape job.

    Returns ``{lookup, lookup_compact, lookup_by_url, feed_rows}``.
    """
    xlsx_path = fetch_vevor_feed(VEVOR_AU_FEED_URL)
    try:
        lookup, lookup_compact, pos_rows = load_veror_via_excel_positions(xlsx_path)
    finally:
        try:
            os.unlink(xlsx_path)
        except OSError:
            pass
    lookup_by_url: dict[str, dict] = {}
    for entry in lookup.values():
        link = (entry.get("Product Link") or "").strip()
        if not link:
            continue
        key = normalize_vevor_product_url(link)
        if key:
            lookup_by_url[key] = entry
    return {
        "lookup": lookup,
        "lookup_compact": lookup_compact,
        "lookup_by_url": lookup_by_url,
        "feed_rows": pos_rows,
    }


def clean_id(value) -> str:
    """Normalize a SKU cell: strip whitespace, zero-width chars, trailing .0 from Excel floats."""
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\u200b", "").replace("\xa0", " ").strip()
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


_COMPACT_RE = re.compile(r"[^A-Za-z0-9]+")


def compact_id(sku: str) -> str:
    """Return an alphanumeric-only lowercase key for fuzzy SKU matches."""
    if not sku:
        return ""
    return _COMPACT_RE.sub("", sku).lower()


def parse_price_value(value) -> float:
    """Accept 'USD 12.34', '12,34', '12.34', 12.34 — return float (0.0 on failure)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float, Decimal)):
        try:
            return float(value)
        except Exception:
            return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("\xa0", " ").strip()
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except Exception:
        return 0.0


def parse_inventory_value(value) -> int:
    """Accept '10', '10 units', '', None — return int >= 0."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        try:
            n = int(value)
            return max(0, n)
        except Exception:
            return 0
    s = str(value).strip()
    if not s:
        return 0
    m = re.search(r"\d+", s.replace(",", ""))
    if not m:
        return 0
    try:
        return max(0, int(m.group(0)))
    except Exception:
        return 0


def round_precise(val: float, digits: int = 2) -> float:
    try:
        return float(Decimal(str(val)).quantize(Decimal("1." + ("0" * digits))))
    except Exception:
        return round(float(val), digits)


# Legacy positional layout (pre-2026 feed): A=SKU, G=Price, I=Inventory.
LEGACY_SKU_COL = 0
LEGACY_PRICE_COL = 6
LEGACY_INVENTORY_COL = 8


def _normalize_header(value) -> str:
    """Lowercase a header cell and collapse whitespace for tolerant matching."""
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def resolve_vevor_feed_columns(header_row) -> tuple[int, int, int, str]:
    """
    Map the feed header row to ``(sku_idx, price_idx, inventory_idx, mode)``.

    The live Vevor AU feed uses ``after coupon price`` (Excel column 35 /
    0-based index 34) as the sellable cost, with ``MAP (Minimum Advertised
    Price)`` beside it. Stock is ``Inventory quantity``. Column G is
    ``Availability`` and column I is product weight, so the legacy
    positional read yields $0 prices and bogus stock.

    ``mode`` is ``"header"`` when all three columns were found by name, else
    ``"legacy"`` (positional fallback for the old A/G/I layout).
    """
    sku_idx = price_idx = inventory_idx = None
    map_price_idx = None
    for idx, cell in enumerate(header_row or ()):
        name = _normalize_header(cell)
        if not name:
            continue
        if sku_idx is None and name == "sku":
            sku_idx = idx
        elif price_idx is None and (
            name == "after coupon price"
            or ("after" in name and "coupon" in name and "price" in name)
        ):
            price_idx = idx
        elif map_price_idx is None and (
            name == "map (minimum advertised price)"
            or ("map" in name and "price" in name)
            or name in ("price", "posted price")
        ):
            map_price_idx = idx
        elif inventory_idx is None and name in (
            "inventory quantity",
            "inventory",
            "posted inventory",
        ):
            inventory_idx = idx
    if price_idx is None:
        price_idx = map_price_idx
    if sku_idx is not None and price_idx is not None and inventory_idx is not None:
        return sku_idx, price_idx, inventory_idx, "header"
    return LEGACY_SKU_COL, LEGACY_PRICE_COL, LEGACY_INVENTORY_COL, "legacy"


def _cell(row, idx):
    """Safe positional access — read_only rows omit trailing empty cells."""
    return row[idx] if idx < len(row) else None


def load_veror_via_excel_positions(path: str) -> tuple[dict, dict, int]:
    """
    Read the Vevor AU feed XLSX into SKU lookups.

    Columns are resolved from the header row via ``resolve_vevor_feed_columns``
    (SKU / after coupon price / Inventory quantity), falling back to the legacy
    positional layout A=0 SKU, G=6 Price, I=8 Inventory when the header is not
    recognizable.

    Returns ``(lookup, lookup_compact, pos_rows_scanned)``:
    - ``lookup``: {sku: {'Posted Price': float, 'Posted Inventory': int}}
    - ``lookup_compact``: same by ``compact_id(sku)`` for fuzzy match fallback.
    - ``pos_rows_scanned``: number of data rows scanned.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        lookup: dict[str, dict] = {}
        lookup_compact: dict[str, dict] = {}
        pos_rows = 0
        sku_idx = price_idx = inventory_idx = None
        link_idx = None
        mode = "legacy"
        priced_rows = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                sku_idx, price_idx, inventory_idx, mode = resolve_vevor_feed_columns(row)
                link_idx = _product_link_column_index(row)
                if mode == "legacy":
                    logger.warning(
                        "Vevor AU feed header not recognized (%r...); using legacy "
                        "positional columns A/G/I — prices may be wrong if the feed "
                        "layout changed.",
                        list(row or ())[:5],
                    )
                continue
            if row is None or not any(cell is not None for cell in row):
                continue
            pos_rows += 1
            sku = clean_id(_cell(row, sku_idx))
            if not sku:
                continue
            price = round_precise(parse_price_value(_cell(row, price_idx)), 2)
            stock = parse_inventory_value(_cell(row, inventory_idx))
            if price > 0:
                priced_rows += 1
            entry = {"Posted Price": price, "Posted Inventory": int(stock)}
            if link_idx is not None:
                link_raw = _cell(row, link_idx)
                link = str(link_raw).strip() if link_raw is not None else ""
                if link:
                    entry["Product Link"] = link
            lookup[sku] = entry
            ckey = compact_id(sku)
            if ckey:
                lookup_compact[ckey] = entry
        if lookup and priced_rows == 0:
            logger.warning(
                "Vevor AU feed parsed %s SKUs but every price is 0 "
                "(mode=%s, price column index=%s) — feed layout may have changed.",
                len(lookup), mode, price_idx,
            )
        else:
            logger.info(
                "Vevor AU feed parsed: %s SKUs, %s with price > 0 (mode=%s, "
                "sku=%s price=%s inventory=%s).",
                len(lookup), priced_rows, mode, sku_idx, price_idx, inventory_idx,
            )
        return lookup, lookup_compact, pos_rows
    finally:
        wb.close()


def fetch_vevor_feed(url: str | None = None, timeout: int = 60) -> str:
    """Download the Vevor AU XLSX to a temp file and return its path."""
    feed_url = resolve_vevor_au_feed_url(url)
    if not feed_url.startswith(("http://", "https://")):
        raise ValueError(
            f"VEVOR_AU_FEED_URL must be an http(s) URL; got {feed_url!r}"
        )
    resp = requests.get(feed_url, timeout=timeout, stream=True)
    resp.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(prefix="vevor_au_", suffix=".xlsx", delete=False)
    try:
        for chunk in resp.iter_content(chunk_size=65536):
            if chunk:
                tmp.write(chunk)
    finally:
        tmp.close()
    return tmp.name


def lookup_sku(lookup: dict, lookup_compact: dict, sku: str) -> dict | None:
    """Try exact SKU first, then alphanumeric-only fuzzy key."""
    if not sku:
        return None
    exact = lookup.get(sku)
    if exact:
        return exact
    ckey = compact_id(sku)
    if not ckey:
        return None
    return lookup_compact.get(ckey)


def iter_vevor_entries(lookup: dict) -> Iterable[tuple[str, dict]]:
    return lookup.items()


__all__ = [
    "DEFAULT_VEVOR_AU_FEED_URL",
    "VEVOR_AU_FEED_URL",
    "resolve_vevor_au_feed_url",
    "_ingest_only_result",
    "is_vevor_vendor_code",
    "is_vevor_product_url",
    "normalize_vevor_product_url",
    "vevor_identity_candidates",
    "lookup_vevor_price_stock",
    "load_vevor_feed_lookups",
    "clean_id",
    "compact_id",
    "parse_price_value",
    "parse_inventory_value",
    "round_precise",
    "resolve_vevor_feed_columns",
    "load_veror_via_excel_positions",
    "fetch_vevor_feed",
    "lookup_sku",
]
